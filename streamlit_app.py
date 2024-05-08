import requests
import smtplib
from email.mime.text import MIMEText
from openpyxl import load_workbook
import streamlit as st
import time



def main():
    st.title("Website Change Monitor")
    st.write("Problem Statement is as follows:- Legal Compliances and Product Safety")
    st.write("Here we are creating a Website Change Monitor system to automate tracking, promptly detect modifications, and efficiently alert users or administrators. This solution aims to enhance efficiency, save time, and keep users informed about changes across designated websites.")

    st.sidebar.header("Email Configuration")
    sender_email = st.sidebar.text_input("Sender Email")
    password = st.sidebar.text_input("Password", type="password")
    receiver_email = st.sidebar.text_input("Receiver Email")


    filename = "Websites.xlsx"

    # Upload the Excel file
    with open(filename, "rb") as file:
        websites_file = st.file_uploader("Upload Excel file containing websites", file.getvalue(), type=["xlsx"])

    if websites_file is not None:
        # Process the uploaded file
        st.write("File uploaded successfully!")
    if sender_email and password and receiver_email and websites_file:
        try:
            workbook = load_workbook(websites_file)
            worksheet = workbook["Websites"]
            websites = [row[0] for row in worksheet.iter_rows(values_only=True)]

            previous_content = {}

            smtp_server = 'smtp.mail.yahoo.com'
            smtp_port = 465

            st.text("Monitoring websites...")

            while True:
                check_websites(sender_email, password, receiver_email, websites, previous_content, smtp_server,
                               smtp_port)
                time.sleep(300)  # Check every 5 minutes
        except Exception as e:
            st.error(f"Error occurred: {e}")


def send_email(sender_email, password, receiver_email, website):
    subject = f'Website {website} has changed!'
    body = f'The content of {website} has changed.'
    message = MIMEText(body)
    message['Subject'] = subject
    message['From'] = sender_email
    message['To'] = receiver_email
    try:
        with smtplib.SMTP('smtp.mail.yahoo.com', 465) as server:
            server.starttls()
            server.login(sender_email, password)
            server.sendmail(sender_email, receiver_email, message.as_string())
            server.quit()
        st.write(f"Email notification sent for {website}.")

    except Exception as e:
        st.error(f"Failed to send email notification for {website}. Error: {e}")


def check_websites(sender_email, password, receiver_email, websites, previous_content, smtp_server, smtp_port):
    for website in websites:
        response = requests.get(website)
        content = response.text

        if website not in previous_content:
            previous_content[website] = content
        elif previous_content[website] == content:
            st.text(f"No Changes in website:{website}")
        elif previous_content[website] != content:
            send_email(sender_email, password, receiver_email, website)
            previous_content[website] = content


        else:
            st.write(f"Error occurred while checking {website}")


if __name__ == "__main__":
    main()
